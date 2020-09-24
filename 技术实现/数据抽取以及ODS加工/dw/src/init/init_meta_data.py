#!/usr/bin/python
# -*- coding: utf-8 -*-
'''
 脚本名称：init_meta_data.py
 脚本功能：读取excel并将配置数据写入到mysql的元表
 输入参数：excel_name sheet_name condition is_exec
 执行示例：python init_meta_data.py /etl/dw/template/meta/etl_meta_init.xlsx etl_meta_args 1=1 1
 编写人：  guxn
 编写日期：20191107
 修改记录：
 by guxn 20191107  1.新建
'''

import sys
import importlib
import setting
import operate_mysql
import operate_os
import operate_log
from openpyxl import load_workbook

class Init_meta_data:
  def __init__(self,logger,opos,db):
    self.logger = logger
    self.opos= opos
    self.db = db
	
  def get_tab_cols(self,tab_name):
    '''获取表的字段结构'''
    self.logger.info("调用方法 - get_tab_cols")
    try:
        cols=[]
        result=()
        pre_sql="select column_name from information_schema.columns where table_name= '"+tab_name+"' order by ordinal_position ;"
        #获取连接
        result = self.db.search(pre_sql)
        for row in result:
        	cols.append(row[0].__str__())
        self.logger.info("获取表"+tab_name+"字段结构列表成功,字段列表："+str(cols))
    except Exception as err:
        self.logger.error("获取表"+tab_name+"字段结构列表失败."+str(err.args))
    return cols


  def init_meta_data(self,file,tab_name,del_condition):
    '''主程序，实现数据初始化操作'''
    self.logger.info("调用方法 - init_meta_data")
    #将ddl语句写入到文件中
    file_name = setting.SQL_PATH+"meta_data.sql"
    #删除已存在文件
    #opos=operate_os.Operate_os(self.logger)
    self.opos.delete_file(file_name)
    #读取文档
    wb = load_workbook(filename=file, data_only=True)
    ws = wb[tab_name]
    #获取字段结构列表
    cols=self.get_tab_cols(tab_name) 
    #获取字段个数
    num=len(cols)
    self.logger.info( "表 "+tab_name+" 开始组装数据初始化语句...... ")
    #拼接字段语句
    cols_txt=','.join(cols)
    sql_txt="insert into "+tab_name+"  values ("
    #获取语句
    #从第二行开始读取数据
    datas=[]
    for i in range(2,ws.max_row+1):
        data=[]
        for j in range(1,num+1):
            if ws.cell(column=j,row=i).value == None:
                val='Null'
            elif isinstance(ws.cell(column=j,row=i).value,str):
                val="'"+ws.cell(column=j,row=i).value+"'"
            else:
                val=str(ws.cell(column=j,row=i).value)
            data.append(val)
        data_txt=','.join(data) 
        #拼接sql结尾
        sql=sql_txt+data_txt+");"
        #完成一个sql后将sql加入到sql集
        datas.append(sql)
    #执行SQL
    try:
        #先删除已存在的数据，支持按条件删除
        del_sql="delete from "+tab_name+" where "+del_condition+";"
        self.logger.debug( "删除语句："+del_sql)
        self.db.execute_sql(del_sql)
        self.db.commit()
        self.logger.info( "表 "+tab_name+" 存在,数据删除成功. ")
        for txt in datas:
            #将所有sql写入文件
            with open(file_name,'a+') as f:
            	f.write(txt+"\n")
            self.logger.info("写入完成,所有sql已写入文件:"+file_name)
            self.logger.debug( "执行sql:"+txt+"完成.")
    except Exception as err:
        self.logger.error("表"+tab_name+"数据初始化插入时失败,"+str(str(err.args)))
    #finally:
        #self.db.close_mysql()
    self.logger.info( "表"+tab_name+"数据初始化插入成功. ")
    return file_name

  def main(self):
    #判断参数
    if len(sys.argv)<5:
        infostr = "Usage: %s excel tab_name del_condition is_exec" %sys.argv[0]
        self.logger.info("输入参数个数有误，请检查.")
        sys.exit(1)
    self.logger.info("参数个数正确，开始元表数据初始化. ")
    #参数赋值
    #数据库文件名称\excel文件名称\sheet页名称\
    excel_name,sheet_name,del_condition,is_exec = sys.argv[1:] 
    self.logger.info("输入参数为 excel_name:"+excel_name+" sheet_name:"+sheet_name+" del_condition:"+del_condition)
    #执行数据初始化
    sql_file=self.init_meta_data(excel_name,sheet_name,del_condition)
    #选择是否执行
    if int(is_exec) == 1:
        re = self.opos.compile_mysql_file(sql_file)
        if re == 1 :
            print("库表创建编译已完成，请检查执行结果.")
        else:
            print("文件编译出现错误.")
    else:
        pass
    sys.exit(0)

if __name__ == '__main__':
    logger=operate_log.getLogger("/Users/wangdabin1216/git/dw/log/init/init_meta_data.log")
    imt=Init_meta_data(logger,operate_os.Operate_os(logger),operate_mysql.Operate_mysql(logger))
    imt.main()

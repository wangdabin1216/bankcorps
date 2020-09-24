#!/usr/bin/python
# -*- coding: utf-8 -*-
'''
 脚本名称：init_meta_struct.py
 脚本功能：读取excel表结构并初始化mysql中元表结构，可使用新sheet页来实现增量更新处理
 输入参数：excel_name sheet_name is_exec
          python init_meta_struct.py /etl/dw/template/meta/etl_meta_init.xlsx etl_meta_tables 1
 编写人：  guxn
 编写日期：20191107
 修改记录：
 by guxn 20191107  1.新建
'''

import sys
import importlib
import setting
import operate_os
import operate_log
from openpyxl import load_workbook

class Init_meta_struct:
  def __init__(self,logger,opos):
    self.logger = logger
    self.opos= opos

  def get_init_sql(self,cols):
    '''拼接创建语句sql'''
    self.logger.info("方法调用 - get_init_sql")
    tab_name=cols[0][0]
    txt=[]
    key=[]
    #获取创建表语句，带主键
    for col in cols:
        txt.append(str(col[1])+" "+str(col[2]))
        if col[6]==1:
            key.append(col[1])
        else:
            pass
    txts=','.join(txt)
    keys=','.join(key)
    #创建有主键和无主键
    if keys!="":
        sql ="create table "+str(tab_name)+"( \n "+txts+", \n"+ "CONSTRAINT pk_"+str(tab_name)+" PRIMARY KEY ("+keys+")); \n\n"
    else:
        sql="create table "+str(tab_name)+"( \n "+txts+" \n ); \n\n"
    self.logger.debug("sql:"+sql)
    return sql

  def init_meta_struct(self,file,tab_name):
    '''主过程，实现元数据库结构初始化'''
    self.logger.info("方法调用 - init_meta_struct")
    #将ddl语句写入到文件中
    file_name = setting.SQL_PATH+"meta_struct.sql"
    #删除已存在文件
    self.opos.delete_file(file_name)
    
    #读取文档
    wb = load_workbook(filename=file, data_only=True)
    ws = wb[tab_name]
    #获取去重的待初始化表清单
    tables=[]
    for i in range(2,ws.max_row+1):
        if ws.cell(column=1,row=i).value not in tables:
            tables.append(ws.cell(column=1,row=i).value)
        else:
            pass
    
    #每个表
    for table in tables:
        cols=[]
        for row in ws.rows:
            if row[0].value == table:
                line = [col.value for col in row]
                cols.append(line)
            else:
                pass
        #获取语句
        sql=self.get_init_sql(cols)
        #循环写入每个表的DDL
        try:
            #将所有sql写入文件
            with open(file_name,'a+') as f:
            	f.write("drop table if exists "+str(table)+";\n")
            	f.write(sql)
            self.logger.info("表DDL已写入文件:"+file_name)
        except Exception as err:
            self.logger.error("表DDL写入文件时出现错误,错误信息："+str(err.args))
            sys.exit(-1)
    self.logger.info( "所有待初始化元表DDL写入成功. ")
    return file_name

  def main(self):
    #检查输入参数个数
    if len(sys.argv)<4:
        infostr = "Usage: %s excel_name sheet_name,is_exec" %sys.argv[0]
        self.logger.error("输入参数个数有误，请检查.")
        sys.exit(-1)
    self.logger.info( "参数个数正确，开始元表结构初始化. ")
    #获取参数
    excel_name,sheet_name,is_exec = sys.argv[1:] 
    self.logger.info("输入参数为，excel_name:"+excel_name+"sheet_name:"+sheet_name)
    sql_file=self.init_meta_struct(excel_name,sheet_name)
	#选择是否执行
    if int(is_exec) == 1:
        re = self.opos.compile_mysql_file(sql_file)
        if re == 1 :
            print("库表创建编译已完成，请检查执行结果.")
        else:
            print("文件编译出现错误.")
    else:
        pass
    sys.exit(0)

if __name__ == '__main__':
    logger=operate_log.getLogger("/Users/wangdabin1216/git/dw/log/init/init_meta_struct.log")
    imt=Init_meta_struct(logger,operate_os.Operate_os(logger))
    imt.main()
    

